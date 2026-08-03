from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os
import requests
import json
import time
from bs4 import BeautifulSoup

# =====================【配置区】=====================
BILIBILI_UID = "628596718"
EXCEL_PATH = "bili_data.xlsx"
HTML_OUTPUT = "report.html"
# Github Action 运行时不要填写本地Cookie，匿名访问即可
COOKIE_STR = ""
# ====================================================

def fetch_page_data(uid):
    """
    抓取移动端个人主页，网页解析获取粉丝、播放量，替代不稳定官方API
    移动端链接：https://m.bilibili.com/space/UID
    """
    session = requests.Session()
    headers = {
        "User-Agent": "Mozilla/5.0 (Linux; Android 10) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Referer": "https://m.bilibili.com/"
    }
    if COOKIE_STR.strip():
        headers["Cookie"] = COOKIE_STR.strip()

    url = f"https://m.bilibili.com/space/{uid}"
    try:
        resp = session.get(url, headers=headers, timeout=20)
        resp.raise_for_status()
        html = resp.text
        soup = BeautifulSoup(html, "html.parser")

        # 页面内window.__INITIAL_STATE__ 包含所有UP数据
        import re
        match = re.search(r"window\.__INITIAL_STATE__\s*=\s*(\{.+?\});\(function", html, re.S)
        if not match:
            print("页面未找到初始化数据")
            return None, None
        raw_json_str = match.group(1)
        data = json.loads(raw_json_str)

        up_info = data.get("space", {}).get("upInfo", {})
        follower = up_info.get("fans", 0)
        total_view = up_info.get("archiveView", 0)
        print(f"解析成功｜粉丝：{follower} 总播放：{total_view}")
        return int(total_view), int(follower)

    except Exception as e:
        print(f"页面抓取/解析异常: {str(e)}")
        return None, None


def init_excel(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "UP主数据总览"
    header = ["日期", "总播放量", "总粉丝数", "播放环比增长率%", "粉丝环比增长率%"]
    for col, text in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    wb.save(file_path)


def append_data(file_path, now_view, now_fans):
    wb = load_workbook(file_path)
    ws = wb["UP主数据总览"]
    max_row = ws.max_row
    today_str = datetime.now().strftime("%Y-%m-%d")

    if max_row <= 1:
        view_rate = 0.00
        fans_rate = 0.00
    else:
        last_view = ws.cell(row=max_row, column=2).value
        last_fans = ws.cell(row=max_row, column=3).value

        view_rate = round(((now_view - last_view) / last_view) * 100, 2) if last_view else 0
        fans_rate = round(((now_fans - last_fans) / last_fans) * 100, 2) if last_fans else 0

    new_row_num = max_row + 1
    ws.cell(new_row_num, 1, today_str)
    ws.cell(new_row_num, 2, now_view)
    ws.cell(new_row_num, 3, now_fans)
    ws.cell(new_row_num, 4, view_rate)
    ws.cell(new_row_num, 5, fans_rate)
    wb.save(file_path)


def load_all_data(file_path):
    wb = load_workbook(file_path)
    ws = wb["UP主数据总览"]
    data_list = []
    for row in range(2, ws.max_row + 1):
        date_str = ws.cell(row, 1).value
        view = ws.cell(row, 2).value
        fans = ws.cell(row, 3).value
        rate_view = ws.cell(row, 4).value
        rate_fans = ws.cell(row, 5).value
        data_list.append({
            "date": date_str,
            "month": date_str[:7],
            "view": view,
            "fans": fans,
            "rate_view": rate_view,
            "rate_fans": rate_fans
        })
    return data_list


def build_html(data_list, save_path):
    months = sorted(list({d["month"] for d in data_list}), reverse=True)
    data_json = json.dumps(data_list, ensure_ascii=False)
    month_json = json.dumps(months, ensure_ascii=False)

    html_template = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>B站UP主月度数据报表</title>
    <style>
        body {font-family: "Microsoft YaHei", sans-serif; max-width:1260px; margin:30px auto; padding:0 20px;}
        .header-bar {margin-bottom:24px; display:flex; align-items:center; gap:12px}
        select {padding:6px 12px; font-size:16px;}
        .chart-box {margin:40px 0;}
        svg {border:1px solid #ddd; background:#fff; width:100%; height:420px;}
        .legend span {display:inline-block; padding:2px 10px; margin:0 4px; color:#fff;}
    </style>
</head>
<body>
    <h1>B站UP主运营数据看板</h1>
    <div class="header-bar">
        <label>选择月份：</label>
        <select id="monthSelector"></select>
    </div>

    <div class="chart-box">
        <h3>总播放量 & 总粉丝 累计趋势</h3>
        <div class="legend">
            <span style="background:#fb7299">总播放量</span>
            <span style="background:#3498db">总粉丝数</span>
        </div>
        <div id="lineContainer"></div>
    </div>

    <div class="chart-box">
        <h3>环比增长率（%）</h3>
        <div class="legend">
            <span style="background:#fb7299">播放环比增长率</span>
            <span style="background:#3498db">粉丝环比增长率</span>
        </div>
        <div id="barContainer"></div>
    </div>

<script>
const rawData = ''' + data_json + ''';
const monthList = ''' + month_json + ''';
const selector = document.getElementById("monthSelector");

monthList.forEach(m=>{
    const opt = document.createElement("option");
    opt.value = m;
    opt.textContent = m;
    selector.appendChild(opt);
})

const PADDING = {top:40, left:60, right:30, bottom:60};

function drawLineChart(container, dataset) {
    const width = container.clientWidth;
    const height = 420;
    const innerW = width - PADDING.left - PADDING.right;
    const innerH = height - PADDING.top - PADDING.bottom;

    const dates = dataset.map(d=>d.date);
    const views = dataset.map(d=>d.view);
    const fans = dataset.map(d=>d.fans);
    const maxVal = Math.max(...views, ...fans);

    let svg = `<svg width="${width}" height="${height}" xmlns="http://www.w3.org/2000/svg">`;
    svg += `<line x1="${PADDING.left}" y1="${PADDING.top}" x2="${PADDING.left}" y2="${height-PADDING.bottom}" stroke="#888"/>`;
    svg += `<line x1="${PADDING.left}" y1="${height-PADDING.bottom}" x2="${width-PADDING.right}" y2="${height-PADDING.bottom}" stroke="#888"/>`;

    const stepX = innerW/(dates.length-1 || 1);

    let pathV = `M`;
    let pathF = `M`;
    for(let i=0;i<dataset.length;i++){
        const x = PADDING.left + i*stepX;
        const yV = PADDING.top + innerH*(1 - views[i]/maxVal);
        const yF = PADDING.top + innerH*(1 - fans[i]/maxVal);
        pathV += `${x},${yV} `;
        pathF += `${x},${yF} `;
        svg += `<text x="${x}" y="${height-PADDING.bottom+18}" font-size="11" text-anchor="middle">${dates[i].slice(5)}</text>`;
    }
    svg += `<path d="${pathV}" fill="none" stroke="#fb7299" stroke-width="2"/>`;
    svg += `<path d="${pathF}" fill="none" stroke="#3498db" stroke-width="2"/>`;
    svg += `</svg>`;
    container.innerHTML = svg;
}

function drawBarChart(container, dataset){
    const width = container.clientWidth;
    const height = 420;
    const innerW = width - PADDING.left - PADDING.right;
    const innerH = height - PADDING.top - PADDING.bottom;

    const rateV = dataset.map(d=>d.rate_view);
    const rateF = dataset.map(d=>d.rate_fans);
    const allRates = [...rateV, ...rateF];
    const maxRate = Math.max(...allRates, 1);
    const minRate = Math.min(...allRates, -1);
    const range = maxRate - minRate || 1;

    let svg = `<svg width="${width}" height="${height}" xmlns="http://www.w3.org/2000/svg">`;
    svg += `<line x1="${PADDING.left}" y1="${PADDING.top}" x2="${PADDING.left}" y2="${height-PADDING.bottom}" stroke="#888"/>`;
    svg += `<line x1="${PADDING.left}" y1="${height-PADDING.bottom}" x2="${width-PADDING.right}" y2="${height-PADDING.bottom}" stroke="#888"/>`;

    const count = dataset.length;
    const barW = innerW / count / 2.4;
    const zeroY = PADDING.top + innerH * (0 - minRate)/range;

    for(let i=0;i<count;i++){
        const baseX = PADDING.left + i*(innerW/count) + barW*0.2;
        const rv = rateV[i];
        const rf = rateF[i];

        const y1 = PADDING.top + innerH*(rv-minRate)/range;
        const h1 = rv >=0 ? zeroY - y1 : y1-zeroY;
        svg += `<rect x="${baseX}" y="${Math.min(y1,zeroY)}" width="${barW}" height="${Math.abs(h1)}" fill="#fb729980"/>`;

        const y2 = PADDING.top + innerH*(rf-minRate)/range;
        const h2 = rf >=0 ? zeroY - y2 : y2-zeroY;
        svg += `<rect x="${baseX+barW+4}" y="${Math.min(y2,zeroY)}" width="${barW}" height="${Math.abs(h2)}" fill="#3498db80"/>`;

        svg += `<text x="${baseX+barW}" y="${height-PADDING.bottom+18}" font-size="11" text-anchor="middle">${dataset[i].date.slice(5)}</text>`;
    }
    svg += `<line x1="${PADDING.left}" y1="${zeroY}" x2="${width-PADDING.right}" stroke="#aaa" stroke-dasharray="4"/>`;
    svg += `</svg>`;
    container.innerHTML = svg;
}

function render(month){
    const filter = rawData.filter(item=>item.month === month);
    filter.sort((a,b)=>a.date.localeCompare(b.date));
    if(filter.length === 0) return;
    drawLineChart(document.getElementById("lineContainer"), filter);
    drawBarChart(document.getElementById("barContainer"), filter);
}

selector.onchange = ()=>render(selector.value);
render(monthList[0]);
</script>
</body>
</html>
'''
    with open(save_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"HTML已生成: {save_path}")


def main():
    total_view, total_fans = fetch_page_data(BILIBILI_UID)
    if total_view is None or total_fans is None:
        print("❌ 数据抓取失败，程序退出")
        return

    if not os.path.exists(EXCEL_PATH):
        init_excel(EXCEL_PATH)
    append_data(EXCEL_PATH, total_view, total_fans)
    all_data = load_all_data(EXCEL_PATH)
    build_html(all_data, HTML_OUTPUT)
    print("✅ 全部任务执行完成")


if __name__ == "__main__":
    main()
